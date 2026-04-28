# -*- coding: utf-8 -*-
from __future__ import annotations

import base64
import hashlib
import html
from dataclasses import dataclass
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


SHEET_NAME = "Лист1"
ROW_LABEL_KEY = "__row_label__"
FONT_CACHE: dict[tuple[int, bool], ImageFont.ImageFont] = {}
FISH_ASSET_FILES = {
    "evil": "bad_fish.png",
    "good": "good_fish.png",
}
DEFAULT_LAYOUT = {
    "row_width": 650,
    "cell_width": 260,
}


@dataclass(frozen=True)
class GroupRange:
    key: str
    label: str
    columns: tuple[int, ...]


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_table_value(value: Any) -> str:
    return clean_cell(value).replace("≈", "-")


def excel_col_key(col_index: int) -> str:
    return f"c_{col_index}"


def is_arrow_like(value: str) -> bool:
    if not value:
        return False
    return all(char in "↑↓≈~—-–/ \\t" for char in value)


def has_group_data(ws: Any, col_index: int) -> bool:
    return any(clean_cell(ws.cell(row=row, column=col_index).value) for row in range(2, ws.max_row + 1))


def find_group_ranges(ws: Any) -> list[GroupRange]:
    groups: list[GroupRange] = []

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == 1 and merged_range.max_row == 1 and merged_range.min_col > 1:
            label = clean_cell(ws.cell(row=1, column=merged_range.min_col).value)
            if not label:
                continue
            columns = tuple(
                col
                for col in range(merged_range.min_col, merged_range.max_col + 1)
                if has_group_data(ws, col)
            )
            if columns:
                groups.append(GroupRange(f"g_{len(groups)}", label, columns))

    if groups:
        return sorted(groups, key=lambda group: group.columns[0])

    header_cells = [
        (col, clean_cell(ws.cell(row=1, column=col).value))
        for col in range(2, ws.max_column + 1)
        if clean_cell(ws.cell(row=1, column=col).value)
    ]
    for index, (start_col, label) in enumerate(header_cells):
        next_start = header_cells[index + 1][0] if index + 1 < len(header_cells) else ws.max_column + 2
        candidate_columns = range(start_col, next_start - 1)
        columns = tuple(col for col in candidate_columns if has_group_data(ws, col))
        if columns:
            groups.append(GroupRange(f"g_{len(groups)}", label, columns))

    return groups


def find_drug_header_row(ws: Any, groups: list[GroupRange]) -> int:
    data_columns = [col for group in groups for col in group.columns]
    for row in range(2, ws.max_row + 1):
        first_col = clean_cell(ws.cell(row=row, column=1).value)
        values = [clean_cell(ws.cell(row=row, column=col).value) for col in data_columns]
        values = [value for value in values if value]
        if first_col or len(values) < max(2, len(data_columns) // 3):
            continue
        arrow_values = sum(1 for value in values if is_arrow_like(value))
        if arrow_values <= len(values) // 2:
            return row
    raise ValueError("Не удалось найти строку с названиями препаратов на Лист1.")


def non_empty_data_row(ws: Any, row: int, groups: list[GroupRange]) -> bool:
    if clean_cell(ws.cell(row=row, column=1).value):
        return True
    return any(
        clean_cell(ws.cell(row=row, column=col).value)
        for group in groups
        for col in group.columns
    )


def read_section(ws: Any, rows: list[int], group: GroupRange) -> list[dict[str, str]]:
    section: list[dict[str, str]] = []
    for row in rows:
        item: dict[str, str] = {ROW_LABEL_KEY: clean_cell(ws.cell(row=row, column=1).value)}
        for col in group.columns:
            item[excel_col_key(col)] = normalize_table_value(ws.cell(row=row, column=col).value)
        section.append(item)
    return section


@st.cache_data(show_spinner=False)
def parse_workbook(file_bytes: bytes) -> dict[str, Any]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    groups = find_group_ranges(ws)
    if not groups:
        raise ValueError("На первой строке не найдены группы с данными.")

    header_row = find_drug_header_row(ws, groups)
    top_rows = [row for row in range(2, header_row) if non_empty_data_row(ws, row, groups)]
    bottom_rows = [
        row
        for row in range(header_row + 1, ws.max_row + 1)
        if non_empty_data_row(ws, row, groups)
    ]

    parsed_groups: list[dict[str, Any]] = []
    for index, group in enumerate(groups):
        columns = [
            {
                "key": excel_col_key(col),
                "excel_column": get_column_letter(col),
                "label": clean_cell(ws.cell(row=header_row, column=col).value) or f"Колонка {get_column_letter(col)}",
            }
            for col in group.columns
        ]
        parsed_groups.append(
            {
                "key": group.key,
                "label": group.label,
                "fish_mood": "evil" if index == 0 else "good",
                "columns": columns,
                "top": read_section(ws, top_rows, group),
                "bottom": read_section(ws, bottom_rows, group),
            }
        )

    return {
        "sheet": ws.title,
        "header_row": header_row,
        "groups": parsed_groups,
    }


def source_bytes(uploaded_file: Any | None) -> tuple[bytes, str]:
    if uploaded_file is not None:
        data = uploaded_file.getvalue()
        return data, uploaded_file.name
    st.info("Загрузите `.xlsx` файл через боковую панель, после этого появятся редактор, визуализация и экспорт PNG.")
    st.stop()


def make_editor_frame(rows: list[dict[str, str]], columns: list[dict[str, str]]) -> pd.DataFrame:
    ordered_columns = [ROW_LABEL_KEY] + [column["key"] for column in columns]
    frame = pd.DataFrame(rows, columns=ordered_columns).fillna("")
    return frame.astype(str)


def rows_from_frame(frame: pd.DataFrame, columns: list[dict[str, str]]) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for _, row in frame.fillna("").iterrows():
        label = clean_cell(row.get(ROW_LABEL_KEY, ""))
        if not label and not any(clean_cell(row.get(column["key"], "")) for column in columns):
            continue
        item = {ROW_LABEL_KEY: label}
        for column in columns:
            item[column["key"]] = normalize_table_value(row.get(column["key"], ""))
        records.append(item)
    return records


def cell_class(value: str) -> str:
    has_up = "↑" in value
    has_down = "↓" in value
    if has_up and has_down:
        return "mixed"
    if has_up:
        return "up"
    if has_down:
        return "down"
    if "≈" in value or "~" in value or "—" in value or "-" == value:
        return "neutral"
    return "empty"


def arrow_directions(value: str) -> list[str]:
    return [char for char in value if char in "↑↓"]


def mixed_cell_gradient(value: str) -> str:
    directions = arrow_directions(value)
    if len(directions) < 2:
        return ""

    stops = []
    for index, direction in enumerate(directions):
        css_color = "var(--up)" if direction == "↑" else "var(--down)"
        start = index * 100 / len(directions)
        end = (index + 1) * 100 / len(directions)
        stops.append(f"{css_color} {start:.3f}% {end:.3f}%")
    return f"background: linear-gradient(90deg, {', '.join(stops)});"


def fish_asset_path(mood: str) -> Path:
    mood = "evil" if mood == "evil" else "good"
    return Path(__file__).with_name(FISH_ASSET_FILES[mood])


def remove_light_checker_background(image: Image.Image) -> Image.Image:
    rgba = image.convert("RGBA")
    if rgba.getchannel("A").getextrema() != (255, 255):
        return rgba

    data = []
    for r, g, b, _ in rgba.getdata():
        near_white = r > 232 and g > 232 and b > 232 and max(r, g, b) - min(r, g, b) < 14
        data.append((r, g, b, 0 if near_white else 255))
    rgba.putdata(data)
    return rgba


def crop_to_alpha(image: Image.Image, padding: int = 24, alpha_threshold: int = 5) -> Image.Image:
    rgba = image.convert("RGBA")
    alpha = rgba.getchannel("A")
    mask = alpha.point(lambda value: 255 if value > alpha_threshold else 0)
    bbox = mask.getbbox()
    if not bbox:
        return rgba
    left = max(0, bbox[0] - padding)
    top = max(0, bbox[1] - padding)
    right = min(rgba.width, bbox[2] + padding)
    bottom = min(rgba.height, bbox[3] + padding)
    return rgba.crop((left, top, right, bottom))


@lru_cache(maxsize=4)
def processed_fish_asset_bytes(mood: str) -> bytes:
    path = fish_asset_path(mood)
    if not path.exists():
        return b""
    with Image.open(path) as source:
        image = remove_light_checker_background(source)
        image = crop_to_alpha(image)
        output = BytesIO()
        image.save(output, format="PNG")
        return output.getvalue()


def fish_image_html(mood: str) -> str:
    data = processed_fish_asset_bytes(mood)
    if not data:
        return zebra_fish_svg(mood, hashlib.md5(mood.encode()).hexdigest()[:6])
    encoded = base64.b64encode(data).decode("ascii")
    alt = "Злая зебрафиш" if mood == "evil" else "Добрая зебрафиш"
    return f'<img class="fish-img fish-{html.escape(mood)}" src="data:image/png;base64,{encoded}" alt="{html.escape(alt)}"/>'


def zebra_fish_svg(mood: str, uid: str) -> str:
    mood = "evil" if mood == "evil" else "good"
    if mood == "evil":
        body = "#d9eef5"
        stripe = "#202020"
        fin = "#d84f47"
        eye = "#e03131"
        brow = """
        <path d="M169 44 L188 34" stroke="#111" stroke-width="5" stroke-linecap="round"/>
        <circle cx="181" cy="49" r="7" fill="#e03131"/>
        <circle cx="183" cy="49" r="2.5" fill="#111"/>
        """
        mouth = """
        <path d="M199 68 Q210 59 222 64" fill="none" stroke="#111" stroke-width="4" stroke-linecap="round"/>
        <path d="M210 66 L215 78 L220 66" fill="#fff" stroke="#111" stroke-width="1.5"/>
        """
        dorsal = """
        <path d="M88 20 L101 5 L116 22 L131 7 L145 25" fill="#d84f47" stroke="#111" stroke-width="2"/>
        """
    else:
        body = "#dff8f1"
        stripe = "#1c3f58"
        fin = "#f4bd4f"
        eye = "#17212b"
        brow = """
        <circle cx="181" cy="49" r="7" fill="#fff"/>
        <circle cx="183" cy="49" r="3" fill="#17212b"/>
        <circle cx="185" cy="47" r="1.2" fill="#fff"/>
        <circle cx="169" cy="67" r="5" fill="#ff9aa2" opacity=".65"/>
        """
        mouth = """
        <path d="M199 65 Q211 75 224 64" fill="none" stroke="#17212b" stroke-width="4" stroke-linecap="round"/>
        """
        dorsal = """
        <path d="M92 22 Q119 7 146 24" fill="none" stroke="#f4bd4f" stroke-width="9" stroke-linecap="round"/>
        """

    return f"""
    <svg class="fish-svg fish-{mood}" viewBox="0 0 260 120" role="img" aria-label="{html.escape(mood)} zebrafish">
      <defs>
        <clipPath id="fish-body-{uid}">
          <path d="M56 59 C78 24 139 18 188 38 C218 50 235 69 224 82 C210 101 154 107 97 96 C72 91 56 79 50 67 C45 67 38 70 30 78 L19 91 L23 69 L5 61 L24 54 L18 32 L32 44 C40 51 47 55 56 59 Z"/>
        </clipPath>
      </defs>
      <path d="M55 60 L18 31 L25 56 L5 62 L25 68 L18 92 L55 67 Z" fill="{fin}" stroke="#17212b" stroke-width="3" stroke-linejoin="round"/>
      {dorsal}
      <path d="M56 59 C78 24 139 18 188 38 C218 50 235 69 224 82 C210 101 154 107 97 96 C72 91 56 79 50 67 C45 67 38 70 30 78 L19 91 L23 69 L5 61 L24 54 L18 32 L32 44 C40 51 47 55 56 59 Z" fill="{body}" stroke="#17212b" stroke-width="3" stroke-linejoin="round"/>
      <g clip-path="url(#fish-body-{uid})">
        <path d="M74 24 L93 105" stroke="{stripe}" stroke-width="10" opacity=".95"/>
        <path d="M105 19 L126 105" stroke="{stripe}" stroke-width="10" opacity=".95"/>
        <path d="M139 23 L158 102" stroke="{stripe}" stroke-width="10" opacity=".95"/>
        <path d="M169 33 L184 94" stroke="{stripe}" stroke-width="9" opacity=".95"/>
      </g>
      <path d="M113 88 Q132 108 154 89" fill="{fin}" stroke="#17212b" stroke-width="2"/>
      {brow}
      {mouth}
      <circle cx="181" cy="49" r="8.5" fill="none" stroke="#17212b" stroke-width="2"/>
    </svg>
    """


def render_table_html(
    section_title: str,
    rows: list[dict[str, str]],
    columns: list[dict[str, str]],
    row_label_title: str,
    show_header: bool = True,
) -> str:
    header_cells = "".join(f"<th>{html.escape(column['label'])}</th>" for column in columns)
    table_head = (
        f"""
        <thead>
          <tr><th class="row-head">{html.escape(row_label_title)}</th>{header_cells}</tr>
        </thead>
        """
        if show_header
        else ""
    )
    body_rows = []
    for row in rows:
        label = html.escape(row.get(ROW_LABEL_KEY, ""))
        cells = [f"<th class='row-name'>{label}</th>"]
        for column in columns:
            value = row.get(column["key"], "")
            class_name = cell_class(value)
            style = mixed_cell_gradient(value) if class_name == "mixed" else ""
            style_attr = f' style="{style}"' if style else ""
            cells.append(
                "<td class='arrow-cell "
                + class_name
                + "'"
                + style_attr
                + "><span>"
                + html.escape(value)
                + "</span></td>"
            )
        body_rows.append("<tr>" + "".join(cells) + "</tr>")

    return f"""
    <div class="section-title">{html.escape(section_title)}</div>
    <div class="table-wrap">
      <table class="zebra-table">
        {table_head}
        <tbody>
          {''.join(body_rows)}
        </tbody>
      </table>
    </div>
    """


def render_figure_html(
    groups: list[dict[str, Any]],
    labels: dict[str, str],
    colors: dict[str, str],
) -> str:
    group_html = []
    for index, group in enumerate(groups):
        mood = group.get("fish_mood", "good")
        fish = fish_image_html(mood)
        top_table = render_table_html(labels["top_section"], group["top"], group["columns"], labels["row_label"])
        bottom_table = render_table_html(labels["bottom_section"], group["bottom"], group["columns"], labels["row_label"], show_header=False)
        group_html.append(
            f"""
            <section class="group-panel {html.escape(mood)}">
              <header class="group-header">
                <div class="fish-box">
                  {fish}
                </div>
                <div class="group-title">
                  <h2>{html.escape(group["label"])}</h2>
                </div>
              </header>
              {top_table}
              {bottom_table}
            </section>
            """
        )

    return f"""
    <!doctype html>
    <html lang="ru">
    <head>
      <meta charset="utf-8"/>
      <style>
        :root {{
          --up: {colors["up"]};
          --down: {colors["down"]};
          --neutral: {colors["neutral"]};
          --empty: {colors["empty"]};
          --ink: #17212b;
          --muted: #59646f;
          --line: #c7d0d8;
          --panel: #ffffff;
          --paper: #f5f7f8;
        }}
        * {{ box-sizing: border-box; }}
        body {{
          margin: 0;
          font-family: Inter, Arial, Helvetica, sans-serif;
          color: var(--ink);
          background: var(--paper);
          letter-spacing: 0;
        }}
        .figure {{
          padding: 20px;
          background: var(--paper);
        }}
        .title-row {{
          display: flex;
          justify-content: space-between;
          gap: 16px;
          align-items: end;
          margin-bottom: 16px;
          border-bottom: 3px solid var(--ink);
          padding-bottom: 12px;
        }}
        h1 {{
          margin: 0;
          font-size: 28px;
          line-height: 1.1;
        }}
        .subtitle {{
          margin-top: 6px;
          color: var(--muted);
          font-size: 15px;
          line-height: 1.35;
        }}
        .legend {{
          display: flex;
          flex-wrap: wrap;
          gap: 8px;
          justify-content: flex-end;
          min-width: 320px;
        }}
        .legend-item {{
          display: inline-flex;
          align-items: center;
          gap: 7px;
          padding: 6px 8px;
          border: 1px solid var(--line);
          border-radius: 8px;
          background: #fff;
          font-size: 13px;
          white-space: nowrap;
        }}
        .swatch {{
          width: 20px;
          height: 14px;
          border-radius: 4px;
          border: 1px solid rgba(0,0,0,.16);
        }}
        .swatch.up {{ background: var(--up); }}
        .swatch.down {{ background: var(--down); }}
        .swatch.neutral {{ background: var(--neutral); }}
        .swatch.mixed {{ background: linear-gradient(90deg, var(--down) 0 33.333%, var(--up) 33.333% 66.667%, var(--down) 66.667% 100%); }}
        .groups {{
          display: grid;
          grid-template-columns: repeat(auto-fit, minmax(460px, 1fr));
          gap: 16px;
          align-items: start;
        }}
        .group-panel {{
          background: var(--panel);
          border: 2px solid var(--ink);
          border-radius: 8px;
          overflow: hidden;
        }}
        .group-header {{
          display: grid;
          grid-template-columns: 190px 1fr;
          gap: 12px;
          align-items: center;
          min-height: 132px;
          padding: 14px;
          border-bottom: 2px solid var(--ink);
          background: linear-gradient(90deg, #ffffff 0%, #edf6f9 100%);
        }}
        .group-panel.evil .group-header {{
          background: linear-gradient(90deg, #ffffff 0%, #ffe9e6 100%);
        }}
        .fish-box {{
          min-width: 0;
        }}
        .fish-svg,
        .fish-img {{
          width: 100%;
          max-width: 190px;
          max-height: 106px;
          height: auto;
          object-fit: contain;
          display: block;
        }}
        .group-title h2 {{
          margin: 0;
          font-size: 28px;
          line-height: 1.05;
          overflow-wrap: anywhere;
        }}
        .section-title {{
          padding: 9px 12px;
          background: #17212b;
          color: #fff;
          font-weight: 800;
          font-size: 15px;
          line-height: 1.25;
        }}
        .table-wrap {{
          overflow-x: auto;
          border-bottom: 1px solid var(--line);
        }}
        .zebra-table {{
          width: 100%;
          border-collapse: collapse;
          table-layout: fixed;
          min-width: 680px;
        }}
        .zebra-table th,
        .zebra-table td {{
          border: 1px solid var(--line);
          padding: 7px 6px;
          text-align: center;
          vertical-align: middle;
          font-size: 13px;
          line-height: 1.2;
          overflow-wrap: anywhere;
        }}
        .zebra-table thead th {{
          background: #f0f4f6;
          font-weight: 800;
        }}
        .zebra-table thead th:not(.row-head) {{
          background: #fff0b8;
        }}
        .zebra-table .row-head,
        .zebra-table .row-name {{
          width: 190px;
          text-align: left;
          font-weight: 800;
          background: #f8fafb;
        }}
        .arrow-cell {{
          font-weight: 900;
          font-size: 18px;
          min-width: 58px;
          color: #071018;
        }}
        .arrow-cell.up {{ background: var(--up); }}
        .arrow-cell.down {{ background: var(--down); }}
        .arrow-cell.neutral {{ background: var(--neutral); }}
        .arrow-cell.empty {{ background: var(--empty); color: #8a949e; }}
        .arrow-cell.mixed {{ background: linear-gradient(90deg, var(--up) 0 50%, var(--down) 50% 100%); }}
        .arrow-cell span {{
          display: inline-flex;
          align-items: center;
          justify-content: center;
          min-width: 32px;
          min-height: 24px;
          padding: 1px 4px;
          border-radius: 6px;
          background: rgba(255,255,255,.58);
        }}
        @media (max-width: 760px) {{
          .figure {{ padding: 12px; }}
          .title-row {{ display: block; }}
          .legend {{ justify-content: flex-start; margin-top: 12px; min-width: 0; }}
          .groups {{ grid-template-columns: 1fr; }}
          .group-header {{ grid-template-columns: 1fr; }}
          .fish-svg,
          .fish-img {{ margin: 0 auto; }}
        }}
      </style>
    </head>
    <body>
      <main class="figure">
        <div class="title-row">
          <div>
            <h1>{html.escape(labels["title"])}</h1>
            <div class="subtitle">{html.escape(labels["subtitle"])}</div>
          </div>
          <div class="legend" aria-label="Легенда">
            <div class="legend-item"><span class="swatch up"></span>{html.escape(labels["up_label"])}</div>
            <div class="legend-item"><span class="swatch down"></span>{html.escape(labels["down_label"])}</div>
            <div class="legend-item"><span class="swatch mixed"></span>{html.escape(labels["mixed_label"])}</div>
            <div class="legend-item"><span class="swatch neutral"></span>{html.escape(labels["neutral_label"])}</div>
          </div>
        </div>
        <div class="groups">
          {''.join(group_html)}
        </div>
      </main>
    </body>
    </html>
    """


def estimate_height(groups: list[dict[str, Any]]) -> int:
    max_rows = 0
    for group in groups:
        max_rows = max(max_rows, len(group["top"]) + len(group["bottom"]))
    return max(820, 310 + max_rows * 42)


def hex_to_rgb(value: str) -> tuple[int, int, int]:
    value = value.strip().lstrip("#")
    if len(value) == 3:
        value = "".join(char * 2 for char in value)
    if len(value) != 6:
        return (255, 255, 255)
    return tuple(int(value[index : index + 2], 16) for index in (0, 2, 4))


def get_font(size: int, bold: bool = False) -> ImageFont.ImageFont:
    key = (size, bold)
    if key in FONT_CACHE:
        return FONT_CACHE[key]

    local_font_dir = Path(__file__).with_name("fonts")
    regular_candidates = [
        local_font_dir / "DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeui.ttf",
    ]
    bold_candidates = [
        local_font_dir / "DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
        "C:/Windows/Fonts/segoeuib.ttf",
    ]

    for font_path in bold_candidates if bold else regular_candidates:
        if Path(font_path).exists():
            FONT_CACHE[key] = ImageFont.truetype(str(font_path), size=size)
            return FONT_CACHE[key]

    FONT_CACHE[key] = ImageFont.load_default()
    return FONT_CACHE[key]


def text_bbox(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> tuple[int, int]:
    if not text:
        return (0, 0)
    bbox = draw.textbbox((0, 0), text, font=font)
    return (bbox[2] - bbox[0], bbox[3] - bbox[1])


def wrap_line(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    if text == "":
        return [""]
    words = text.split(" ")
    lines: list[str] = []
    current = ""

    for word in words:
        candidate = word if not current else f"{current} {word}"
        if text_bbox(draw, candidate, font)[0] <= max_width:
            current = candidate
            continue

        if current:
            lines.append(current)
            current = ""

        if text_bbox(draw, word, font)[0] <= max_width:
            current = word
            continue

        chunk = ""
        for char in word:
            candidate_chunk = chunk + char
            if text_bbox(draw, candidate_chunk, font)[0] <= max_width:
                chunk = candidate_chunk
            else:
                if chunk:
                    lines.append(chunk)
                chunk = char
        current = chunk

    if current:
        lines.append(current)
    return lines or [""]


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    lines: list[str] = []
    for line in str(text).splitlines() or [""]:
        lines.extend(wrap_line(draw, line, font, max_width))
    return lines


def line_height(font: ImageFont.ImageFont) -> int:
    bbox = font.getbbox("Ag")
    return bbox[3] - bbox[1]


def wrapped_text_height(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.ImageFont,
    max_width: int,
    spacing: int,
) -> int:
    lines = wrap_text(draw, text, font, max_width)
    return len(lines) * line_height(font) + max(0, len(lines) - 1) * spacing


def draw_wrapped_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    box: tuple[int, int, int, int],
    font: ImageFont.ImageFont,
    fill: tuple[int, int, int],
    align: str = "center",
    valign: str = "center",
    padding: int = 18,
    spacing: int = 8,
) -> None:
    x1, y1, x2, y2 = box
    max_width = max(12, x2 - x1 - padding * 2)
    lines = wrap_text(draw, text, font, max_width)
    text_h = len(lines) * line_height(font) + max(0, len(lines) - 1) * spacing
    if valign == "top":
        y = y1 + padding
    elif valign == "bottom":
        y = y2 - padding - text_h
    else:
        y = y1 + max(0, (y2 - y1 - text_h) // 2)

    for line in lines:
        w, _ = text_bbox(draw, line, font)
        if align == "left":
            x = x1 + padding
        elif align == "right":
            x = x2 - padding - w
        else:
            x = x1 + max(0, (x2 - x1 - w) // 2)
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height(font) + spacing


def cubic_points(
    p0: tuple[float, float],
    p1: tuple[float, float],
    p2: tuple[float, float],
    p3: tuple[float, float],
    steps: int = 18,
) -> list[tuple[float, float]]:
    points = []
    for index in range(1, steps + 1):
        t = index / steps
        mt = 1 - t
        x = mt**3 * p0[0] + 3 * mt**2 * t * p1[0] + 3 * mt * t**2 * p2[0] + t**3 * p3[0]
        y = mt**3 * p0[1] + 3 * mt**2 * t * p1[1] + 3 * mt * t**2 * p2[1] + t**3 * p3[1]
        points.append((x, y))
    return points


def quadratic_points(
    p0: tuple[float, float],
    p1: tuple[float, float],
    p2: tuple[float, float],
    steps: int = 18,
) -> list[tuple[float, float]]:
    points = []
    for index in range(1, steps + 1):
        t = index / steps
        mt = 1 - t
        x = mt**2 * p0[0] + 2 * mt * t * p1[0] + t**2 * p2[0]
        y = mt**2 * p0[1] + 2 * mt * t * p1[1] + t**2 * p2[1]
        points.append((x, y))
    return points


def draw_fish_png(
    image: Image.Image,
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    mood: str,
) -> None:
    x1, y1, x2, y2 = box
    w = x2 - x1
    h = y2 - y1

    def p(px: float, py: float) -> tuple[float, float]:
        return (x1 + w * px / 260, y1 + h * py / 120)

    def pi(px: float, py: float) -> tuple[int, int]:
        sx, sy = p(px, py)
        return (round(sx), round(sy))

    def scale(points: list[tuple[float, float]]) -> list[tuple[int, int]]:
        return [(round(x1 + w * px / 260), round(y1 + h * py / 120)) for px, py in points]

    if mood == "evil":
        body = (216, 238, 245)
        fin = (216, 79, 71)
        stripe = (30, 32, 34)
        eye = (224, 49, 49)
    else:
        body = (223, 248, 241)
        fin = (244, 189, 79)
        stripe = (28, 63, 88)
        eye = (23, 33, 43)

    outline = (23, 33, 43)

    tail = scale([(55, 60), (18, 31), (25, 56), (5, 62), (25, 68), (18, 92), (55, 67)])
    body_curve: list[tuple[float, float]] = [(56, 59)]
    body_curve += cubic_points((56, 59), (78, 24), (139, 18), (188, 38), 20)
    body_curve += cubic_points((188, 38), (218, 50), (235, 69), (224, 82), 16)
    body_curve += cubic_points((224, 82), (210, 101), (154, 107), (97, 96), 20)
    body_curve += cubic_points((97, 96), (72, 91), (56, 79), (50, 67), 14)
    body_curve += cubic_points((50, 67), (45, 67), (38, 70), (30, 78), 8)
    body_curve += [(19, 91), (23, 69), (5, 61), (24, 54), (18, 32), (32, 44)]
    body_curve += cubic_points((32, 44), (40, 51), (47, 55), (56, 59), 10)
    body_poly = scale(body_curve)

    draw.polygon(tail, fill=fin, outline=outline)

    if mood == "evil":
        dorsal_points = scale([(88, 20), (101, 5), (116, 22), (131, 7), (145, 25)])
        draw.polygon(dorsal_points, fill=fin, outline=outline)
    else:
        dorsal = scale([(92, 22)] + quadratic_points((92, 22), (119, 7), (146, 24), 20))
        draw.line(dorsal, fill=fin, width=max(8, int(w * 9 / 260)), joint="curve")

    outline_w = max(3, int(w * 3 / 260))
    draw.polygon(body_poly, fill=body)

    mask = Image.new("L", image.size, 0)
    mask_draw = ImageDraw.Draw(mask)
    mask_draw.polygon(body_poly, fill=255)
    stripes = Image.new("RGBA", image.size, (0, 0, 0, 0))
    stripes_draw = ImageDraw.Draw(stripes)
    for x_start, y_start, x_end, y_end, stroke in (
        (74, 24, 93, 105, 10),
        (105, 19, 126, 105, 10),
        (139, 23, 158, 102, 10),
        (169, 33, 184, 94, 9),
    ):
        stripes_draw.line([pi(x_start, y_start), pi(x_end, y_end)], fill=stripe + (255,), width=max(6, int(w * stroke / 260)))
    image.alpha_composite(Image.composite(stripes, Image.new("RGBA", image.size, (0, 0, 0, 0)), mask))
    draw.line(body_poly + [body_poly[0]], fill=outline, width=outline_w, joint="curve")

    fin_curve = scale([(113, 88)] + quadratic_points((113, 88), (132, 108), (154, 89), 14) + [(113, 88)])
    fin_points = fin_curve
    draw.polygon(fin_points, fill=fin, outline=outline)

    eye_center = pi(181, 49)
    eye_r = max(7, int(w * 0.032))
    if mood == "evil":
        draw.ellipse(
            (eye_center[0] - eye_r, eye_center[1] - eye_r, eye_center[0] + eye_r, eye_center[1] + eye_r),
            fill=eye,
            outline=outline,
            width=3,
        )
        draw.ellipse((eye_center[0] - 3, eye_center[1] - 3, eye_center[0] + 3, eye_center[1] + 3), fill=(17, 17, 17))
        draw.line([pi(169, 44), pi(188, 34)], fill=(17, 17, 17), width=max(4, w // 60))
        mouth = scale([(199, 68)] + quadratic_points((199, 68), (210, 59), (222, 64), 14))
        draw.line(mouth, fill=outline, width=max(3, int(w * 4 / 260)), joint="curve")
        draw.polygon(scale([(210, 66), (215, 78), (220, 66)]), fill=(255, 255, 255), outline=outline)
    else:
        draw.ellipse(
            (eye_center[0] - eye_r, eye_center[1] - eye_r, eye_center[0] + eye_r, eye_center[1] + eye_r),
            fill=(255, 255, 255),
            outline=outline,
            width=3,
        )
        draw.ellipse(
            (eye_center[0] - eye_r // 2, eye_center[1] - eye_r // 2, eye_center[0] + eye_r // 2, eye_center[1] + eye_r // 2),
            fill=eye,
        )
        draw.ellipse((eye_center[0] + 2, eye_center[1] - 4, eye_center[0] + 5, eye_center[1] - 1), fill=(255, 255, 255))
        mouth = scale([(199, 65)] + quadratic_points((199, 65), (211, 75), (224, 64), 14))
        draw.line(mouth, fill=outline, width=max(3, int(w * 4 / 260)), joint="curve")
        cheek_1 = pi(166, 64)
        cheek_2 = pi(176, 73)
        draw.ellipse((cheek_1[0], cheek_1[1], cheek_2[0], cheek_2[1]), fill=(255, 154, 162))

    draw.line([pi(220, 66), pi(238, 65), pi(248, 61)], fill=outline, width=max(2, w // 120))


def paste_fish_asset_png(
    image: Image.Image,
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    mood: str,
) -> None:
    data = processed_fish_asset_bytes(mood)
    if not data:
        draw_fish_png(image, draw, box, mood)
        return

    fish = Image.open(BytesIO(data)).convert("RGBA")
    max_w = box[2] - box[0]
    max_h = box[3] - box[1]
    fish.thumbnail((max_w, max_h), Image.Resampling.LANCZOS)
    x = box[0] + (max_w - fish.width) // 2
    y = box[1] + (max_h - fish.height) // 2
    image.alpha_composite(fish, (x, y))


def mixed_cell_colors(value: str, colors: dict[str, tuple[int, int, int]]) -> list[tuple[int, int, int]]:
    return [colors["up"] if direction == "↑" else colors["down"] for direction in arrow_directions(value)]


def text_width(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> float:
    if not text:
        return 0.0
    if hasattr(draw, "textlength"):
        return float(draw.textlength(text, font=font))
    return float(text_bbox(draw, text, font)[0])


def positioned_text_lines(
    draw: ImageDraw.ImageDraw,
    text: str,
    box: tuple[int, int, int, int],
    font: ImageFont.ImageFont,
    align: str = "center",
    valign: str = "center",
    padding: int = 18,
    spacing: int = 8,
) -> list[tuple[str, float, int]]:
    x1, y1, x2, y2 = box
    max_width = max(12, x2 - x1 - padding * 2)
    lines = wrap_text(draw, text, font, max_width)
    text_h = len(lines) * line_height(font) + max(0, len(lines) - 1) * spacing
    if valign == "top":
        y = y1 + padding
    elif valign == "bottom":
        y = y2 - padding - text_h
    else:
        y = y1 + max(0, (y2 - y1 - text_h) // 2)

    positioned: list[tuple[str, float, int]] = []
    for line in lines:
        w, _ = text_bbox(draw, line, font)
        if align == "left":
            line_x = x1 + padding
        elif align == "right":
            line_x = x2 - padding - w
        else:
            line_x = x1 + max(0, (x2 - x1 - w) // 2)
        positioned.append((line, float(line_x), y))
        y += line_height(font) + spacing
    return positioned


def draw_equal_direction_segments_png(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    segment_colors: list[tuple[int, int, int]],
) -> None:
    x1, y1, x2, y2 = box
    segment_count = max(1, len(segment_colors))
    for index, fill in enumerate(segment_colors):
        left = x1 + round((x2 - x1) * index / segment_count)
        right = x1 + round((x2 - x1) * (index + 1) / segment_count)
        draw.rectangle((left, y1, right, y2), fill=fill)


def draw_aligned_mixed_cell_png(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    value: str,
    colors: dict[str, tuple[int, int, int]],
    font: ImageFont.ImageFont,
) -> None:
    segment_colors = mixed_cell_colors(value, colors)
    if len(segment_colors) < 2:
        draw.rectangle(box, fill=segment_colors[0] if segment_colors else colors["empty"])
        return

    x1, y1, x2, y2 = box
    positioned_lines = positioned_text_lines(draw, value, box, font, padding=10, spacing=2)
    if len(positioned_lines) != 1:
        draw_equal_direction_segments_png(draw, box, segment_colors)
        return

    line, line_x, _ = positioned_lines[0]
    arrow_centers: list[tuple[float, tuple[int, int, int]]] = []
    for char_index, char in enumerate(line):
        if char not in "↑↓":
            continue
        prefix_w = text_width(draw, line[:char_index], font)
        char_w = text_width(draw, line[: char_index + 1], font) - prefix_w
        if char_w <= 0:
            char_w = text_width(draw, char, font)
        fill = colors["up"] if char == "↑" else colors["down"]
        arrow_centers.append((line_x + prefix_w + char_w / 2, fill))

    if len(arrow_centers) != len(segment_colors):
        draw_equal_direction_segments_png(draw, box, segment_colors)
        return

    # Boundaries are placed between rendered arrow centers, so every arrow sits inside its own color band.
    boundaries = [float(x1)]
    boundaries.extend((left[0] + right[0]) / 2 for left, right in zip(arrow_centers, arrow_centers[1:]))
    boundaries.append(float(x2))

    for index, (_, fill) in enumerate(arrow_centers):
        left = max(x1, min(x2, round(boundaries[index])))
        right = max(x1, min(x2, round(boundaries[index + 1])))
        draw.rectangle((left, y1, right, y2), fill=fill)


def cell_fill_png(value: str, colors: dict[str, tuple[int, int, int]]) -> tuple[str, tuple[int, int, int] | None]:
    class_name = cell_class(value)
    if class_name == "up":
        return ("solid", colors["up"])
    if class_name == "down":
        return ("solid", colors["down"])
    if class_name == "neutral":
        return ("solid", colors["neutral"])
    if class_name == "mixed":
        return ("mixed", None)
    return ("solid", colors["empty"])


def normalize_layout(layout: dict[str, int] | None = None) -> dict[str, int]:
    merged = DEFAULT_LAYOUT.copy()
    if layout:
        merged.update({key: int(value) for key, value in layout.items() if key in merged})
    merged["row_width"] = max(360, min(1100, merged["row_width"]))
    merged["cell_width"] = max(150, min(460, merged["cell_width"]))
    return merged


def group_png_width(group: dict[str, Any], layout: dict[str, int]) -> int:
    return layout["row_width"] + len(group["columns"]) * layout["cell_width"]


def measure_table_height(
    draw: ImageDraw.ImageDraw,
    rows: list[dict[str, str]],
    columns: list[dict[str, str]],
    row_label_title: str,
    widths: dict[str, int],
    fonts: dict[str, ImageFont.ImageFont],
    show_header: bool = True,
) -> tuple[int, list[int]]:
    row_heights: list[int] = []
    if show_header:
        header_h = max(
            118,
            max(
                [wrapped_text_height(draw, row_label_title, fonts["table_header"], widths["row"] - 28, 6)]
                + [
                    wrapped_text_height(draw, column["label"], fonts["table_header"], widths["cell"] - 24, 6)
                    for column in columns
                ]
            )
            + 32,
        )
        row_heights.append(header_h)
    for row in rows:
        label_h = wrapped_text_height(draw, row.get(ROW_LABEL_KEY, ""), fonts["row"], widths["row"] - 28, 6)
        cell_h = 0
        for column in columns:
            value = row.get(column["key"], "")
            font = fonts["arrow_small"] if len(value) >= 4 else fonts["arrow"]
            cell_h = max(cell_h, wrapped_text_height(draw, value, font, widths["cell"] - 20, 4))
        row_heights.append(max(96, label_h + 32, cell_h + 26))
    return 78 + sum(row_heights), row_heights


def draw_table_png(
    draw: ImageDraw.ImageDraw,
    x: int,
    y: int,
    width: int,
    section_title: str,
    rows: list[dict[str, str]],
    columns: list[dict[str, str]],
    row_label_title: str,
    colors: dict[str, tuple[int, int, int]],
    fonts: dict[str, ImageFont.ImageFont],
    line: tuple[int, int, int],
    layout: dict[str, int],
    show_header: bool = True,
) -> int:
    widths = {"row": layout["row_width"], "cell": layout["cell_width"]}
    section_h = 78
    _, row_heights = measure_table_height(draw, rows, columns, row_label_title, widths, fonts, show_header=show_header)

    draw.rectangle((x, y, x + width, y + section_h), fill=(23, 33, 43))
    draw_wrapped_text(draw, section_title, (x, y, x + width, y + section_h), fonts["section"], (255, 255, 255), align="left")
    y += section_h

    if show_header:
        header_h = row_heights[0]
        drug_header_fill = (255, 240, 184)
        draw.rectangle((x, y, x + widths["row"], y + header_h), fill=(240, 244, 246), outline=line, width=2)
        draw_wrapped_text(draw, row_label_title, (x, y, x + widths["row"], y + header_h), fonts["table_header"], (23, 33, 43), align="left")
        cx = x + widths["row"]
        for column in columns:
            draw.rectangle((cx, y, cx + widths["cell"], y + header_h), fill=drug_header_fill, outline=line, width=2)
            draw_wrapped_text(draw, column["label"], (cx, y, cx + widths["cell"], y + header_h), fonts["table_header"], (23, 33, 43))
            cx += widths["cell"]
        y += header_h
        data_heights = row_heights[1:]
    else:
        data_heights = row_heights

    for row, row_h in zip(rows, data_heights):
        draw.rectangle((x, y, x + widths["row"], y + row_h), fill=(248, 250, 251), outline=line, width=2)
        draw_wrapped_text(draw, row.get(ROW_LABEL_KEY, ""), (x, y, x + widths["row"], y + row_h), fonts["row"], (23, 33, 43), align="left")

        cx = x + widths["row"]
        for column in columns:
            value = row.get(column["key"], "")
            font = fonts["arrow_small"] if len(value) >= 4 else fonts["arrow"]
            text_w, text_h = text_bbox(draw, value, font)
            badge_w = min(widths["cell"] - 28, max(72, text_w + 28))
            badge_h = min(row_h - 18, max(54, text_h + 22))
            bx = cx + (widths["cell"] - badge_w) // 2
            by = y + (row_h - badge_h) // 2
            mode, solid = cell_fill_png(value, colors)
            if mode == "mixed":
                draw_aligned_mixed_cell_png(draw, (cx, y, cx + widths["cell"], y + row_h), value, colors, font)
            else:
                draw.rectangle((cx, y, cx + widths["cell"], y + row_h), fill=solid or colors["empty"])
            draw.rectangle((cx, y, cx + widths["cell"], y + row_h), outline=line, width=2)
            if value:
                draw.rounded_rectangle((bx, by, bx + badge_w, by + badge_h), radius=12, fill=(255, 255, 255, 185))
            draw_wrapped_text(draw, value, (cx, y, cx + widths["cell"], y + row_h), font, (7, 16, 24), padding=10, spacing=2)
            cx += widths["cell"]
        y += row_h
    return y


def measure_group_height(
    draw: ImageDraw.ImageDraw,
    group: dict[str, Any],
    labels: dict[str, str],
    fonts: dict[str, ImageFont.ImageFont],
    layout: dict[str, int],
) -> int:
    widths = {"row": layout["row_width"], "cell": layout["cell_width"]}
    top_h, _ = measure_table_height(draw, group["top"], group["columns"], labels["row_label"], widths, fonts)
    bottom_h, _ = measure_table_height(draw, group["bottom"], group["columns"], labels["row_label"], widths, fonts, show_header=False)
    return 310 + top_h + bottom_h + 4


def draw_group_png(
    image: Image.Image,
    draw: ImageDraw.ImageDraw,
    x: int,
    y: int,
    width: int,
    height: int,
    group: dict[str, Any],
    labels: dict[str, str],
    colors: dict[str, tuple[int, int, int]],
    fonts: dict[str, ImageFont.ImageFont],
    layout: dict[str, int],
) -> None:
    line = (23, 33, 43)
    header_h = 310
    mood = group.get("fish_mood", "good")
    header_fill = (255, 233, 230) if mood == "evil" else (237, 246, 249)

    draw.rounded_rectangle((x, y, x + width, y + height), radius=18, fill=(255, 255, 255), outline=line, width=5)
    draw.rectangle((x + 3, y + 3, x + width - 3, y + header_h), fill=header_fill)
    draw.line((x, y + header_h, x + width, y + header_h), fill=line, width=5)

    fish_box = (x + 38, y + 36, x + 520, y + 258)
    paste_fish_asset_png(image, draw, fish_box, mood)

    title_x = x + 560
    draw_wrapped_text(
        draw,
        group["label"],
        (title_x, y + 82, x + width - 42, y + header_h - 58),
        fonts["group"],
        line,
        align="left",
        padding=0,
        spacing=10,
    )

    cy = y + header_h
    cy = draw_table_png(draw, x, cy, width, labels["top_section"], group["top"], group["columns"], labels["row_label"], colors, fonts, line, layout)
    draw_table_png(draw, x, cy, width, labels["bottom_section"], group["bottom"], group["columns"], labels["row_label"], colors, fonts, line, layout, show_header=False)


def draw_legend_png(
    draw: ImageDraw.ImageDraw,
    x: int,
    y: int,
    labels: dict[str, str],
    colors: dict[str, tuple[int, int, int]],
    fonts: dict[str, ImageFont.ImageFont],
    width: int,
) -> None:
    items = [
        ("up", labels["up_label"]),
        ("down", labels["down_label"]),
        ("mixed", labels["mixed_label"]),
        ("neutral", labels["neutral_label"]),
    ]
    gap = 14
    iy = y
    for kind, text in items:
        text_h = wrapped_text_height(draw, text, fonts["legend"], width - 120, 6)
        item_h = max(60, text_h + 24)
        draw.rounded_rectangle((x, iy, x + width, iy + item_h), radius=12, fill=(255, 255, 255), outline=(199, 208, 216), width=2)
        swatch_center = iy + item_h // 2
        swatch = (x + 18, swatch_center - 12, x + 58, swatch_center + 12)
        if kind == "mixed":
            third = (swatch[2] - swatch[0]) / 3
            draw.rectangle((swatch[0], swatch[1], round(swatch[0] + third), swatch[3]), fill=colors["down"])
            draw.rectangle((round(swatch[0] + third), swatch[1], round(swatch[0] + third * 2), swatch[3]), fill=colors["up"])
            draw.rectangle((round(swatch[0] + third * 2), swatch[1], swatch[2], swatch[3]), fill=colors["down"])
        else:
            draw.rectangle(swatch, fill=colors[kind])
        draw.rectangle(swatch, outline=(70, 79, 88), width=1)
        draw_wrapped_text(draw, text, (x + 74, iy, x + width - 16, iy + item_h), fonts["legend"], (23, 33, 43), align="left", padding=0)
        iy += item_h + gap


def measure_legend_height(
    draw: ImageDraw.ImageDraw,
    labels: dict[str, str],
    fonts: dict[str, ImageFont.ImageFont],
    width: int,
) -> int:
    items = [labels["up_label"], labels["down_label"], labels["mixed_label"], labels["neutral_label"]]
    heights = [max(60, wrapped_text_height(draw, text, fonts["legend"], width - 120, 6) + 24) for text in items]
    return sum(heights) + 14 * (len(heights) - 1)


def render_png_visualization(
    groups: list[dict[str, Any]],
    labels: dict[str, str],
    color_values: dict[str, str],
    layout: dict[str, int] | None = None,
    dpi: int = 600,
) -> bytes:
    layout = normalize_layout(layout)
    colors = {key: hex_to_rgb(value) for key, value in color_values.items()}
    colors.setdefault("up", (255, 143, 112))
    colors.setdefault("down", (124, 199, 255))
    colors.setdefault("neutral", (232, 237, 240))
    colors.setdefault("empty", (255, 255, 255))
    fonts = {
        "title": get_font(72, True),
        "subtitle": get_font(34),
        "legend": get_font(30, True),
        "group": get_font(64, True),
        "section": get_font(32, True),
        "table_header": get_font(28, True),
        "row": get_font(27, True),
        "arrow": get_font(44, True),
        "arrow_small": get_font(38, True),
    }

    scratch = Image.new("RGBA", (10, 10), (255, 255, 255, 0))
    scratch_draw = ImageDraw.Draw(scratch)
    group_widths = [group_png_width(group, layout) for group in groups]
    group_heights = [measure_group_height(scratch_draw, group, labels, fonts, layout) for group in groups]

    margin = 90
    gap = 90
    legend_w = 780
    title_h = max(330, measure_legend_height(scratch_draw, labels, fonts, legend_w) + 40)
    width = margin * 2 + sum(group_widths) + gap * max(0, len(groups) - 1)
    height = margin + title_h + max(group_heights or [0]) + margin

    image = Image.new("RGBA", (width, height), (245, 247, 248, 255))
    draw = ImageDraw.Draw(image, "RGBA")
    ink = (23, 33, 43)
    muted = (89, 100, 111)

    legend_x = width - margin - legend_w
    title_box = (margin, margin, legend_x - 60, margin + 100)
    subtitle_box = (margin, margin + 112, legend_x - 60, margin + title_h - 32)
    draw_wrapped_text(draw, labels["title"], title_box, fonts["title"], ink, align="left", padding=0, valign="top", spacing=12)
    draw_wrapped_text(draw, labels["subtitle"], subtitle_box, fonts["subtitle"], muted, align="left", padding=0, valign="top", spacing=8)
    draw_legend_png(draw, legend_x, margin, labels, colors, fonts, legend_w)
    draw.line((margin, margin + title_h - 30, width - margin, margin + title_h - 30), fill=ink, width=6)

    x = margin
    y = margin + title_h
    for group, group_width, group_height in zip(groups, group_widths, group_heights):
        draw_group_png(image, draw, x, y, group_width, group_height, group, labels, colors, fonts, layout)
        x += group_width + gap

    output = BytesIO()
    image.convert("RGB").save(output, format="PNG", dpi=(dpi, dpi), optimize=True)
    return output.getvalue()


def default_labels() -> dict[str, str]:

    return {
        "title": "Zebrafish: Neurotoxicity and Neuroactivity",
        "subtitle": "The color of the cell reflects the direction of change; mixed arrows are colored under each arrow.",
        "top_section": "Metabolites and neurotransmitter systems",
        "bottom_section": "Behavioral indicators",
        "row_label": "Parameter",
        "up_label": "Increase: ↑",
        "down_label": "Decrease: ↓",
        "mixed_label": "Mixed directions: ↓↑↓",
        "neutral_label": "No significant change: —",
    }


def sidebar_text_labels(labels: dict[str, str], token: str) -> dict[str, str]:
    edited = labels.copy()
    with st.sidebar.expander("Текстовые подписи", expanded=True):
        edited["title"] = st.text_input("Заголовок", edited["title"], key=f"{token}_title")
        edited["subtitle"] = st.text_area("Подзаголовок", edited["subtitle"], key=f"{token}_subtitle", height=72)
        edited["top_section"] = st.text_input("Раздел сверху", edited["top_section"], key=f"{token}_top_section")
        edited["bottom_section"] = st.text_input("Раздел снизу", edited["bottom_section"], key=f"{token}_bottom_section")
        edited["row_label"] = st.text_input("Заголовок первого столбца", edited["row_label"], key=f"{token}_row_label")
    with st.sidebar.expander("Легенда и цвета", expanded=False):
        edited["up_label"] = st.text_input("Подпись ↑", edited["up_label"], key=f"{token}_up_label")
        edited["down_label"] = st.text_input("Подпись ↓", edited["down_label"], key=f"{token}_down_label")
        edited["mixed_label"] = st.text_input("Подпись смешанных стрелок", edited["mixed_label"], key=f"{token}_mixed_label")
        edited["neutral_label"] = st.text_input("Подпись -", edited["neutral_label"], key=f"{token}_neutral_label")
    return edited


def sidebar_colors(token: str) -> dict[str, str]:
    with st.sidebar.expander("Цвета ячеек", expanded=False):
        up = st.color_picker("↑", "#ff8f70", key=f"{token}_up_color")
        down = st.color_picker("↓", "#7cc7ff", key=f"{token}_down_color")
        neutral = st.color_picker("-", "#e8edf0", key=f"{token}_neutral_color")
        empty = st.color_picker("Пусто", "#ffffff", key=f"{token}_empty_color")
    return {"up": up, "down": down, "neutral": neutral, "empty": empty}


def sidebar_layout_controls(token: str) -> dict[str, int]:
    with st.sidebar.expander("Ширина колонок PNG", expanded=True):
        row_width = st.slider(
            "Первый столбец с показателями",
            min_value=420,
            max_value=950,
            value=DEFAULT_LAYOUT["row_width"],
            step=10,
            key=f"{token}_row_width",
        )
        cell_width = st.slider(
            "Колонки препаратов",
            min_value=180,
            max_value=380,
            value=DEFAULT_LAYOUT["cell_width"],
            step=10,
            key=f"{token}_cell_width",
        )
        st.caption("Эти настройки влияют на PNG-превью и скачиваемую PNG-картинку.")
    return {"row_width": row_width, "cell_width": cell_width}


def edit_group(group: dict[str, Any], labels: dict[str, str], token: str) -> dict[str, Any]:
    group_key = group["key"]
    st.subheader(group["label"])
    group_label = st.text_input("Название группы", group["label"], key=f"{token}_{group_key}_label")

    columns_frame = pd.DataFrame(
        {
            "excel_column": [column["excel_column"] for column in group["columns"]],
            "label": [column["label"] for column in group["columns"]],
        }
    )
    edited_columns_frame = st.data_editor(
        columns_frame,
        key=f"{token}_{group_key}_columns",
        use_container_width=True,
        hide_index=True,
        disabled=["excel_column"],
        column_config={
            "excel_column": st.column_config.TextColumn("Колонка Excel", width="small"),
            "label": st.column_config.TextColumn("Подпись препарата", required=True),
        },
    )
    columns = []
    for original, (_, row) in zip(group["columns"], edited_columns_frame.iterrows()):
        columns.append(
            {
                **original,
                "label": clean_cell(row["label"]) or original["label"],
            }
        )

    column_config = {
        ROW_LABEL_KEY: st.column_config.TextColumn(labels["row_label"], width="medium", required=True),
    }
    column_config.update(
        {
            column["key"]: st.column_config.TextColumn(column["label"], width="small")
            for column in columns
        }
    )

    top_frame = make_editor_frame(group["top"], group["columns"])
    bottom_frame = make_editor_frame(group["bottom"], group["columns"])

    st.markdown(f"**{labels['top_section']}**")
    edited_top = st.data_editor(
        top_frame,
        key=f"{token}_{group_key}_top",
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config=column_config,
    )
    st.markdown(f"**{labels['bottom_section']}**")
    edited_bottom = st.data_editor(
        bottom_frame,
        key=f"{token}_{group_key}_bottom",
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config=column_config,
    )

    return {
        **group,
        "label": group_label,
        "columns": columns,
        "top": rows_from_frame(edited_top, columns),
        "bottom": rows_from_frame(edited_bottom, columns),
    }


def main() -> None:
    st.set_page_config(
        page_title="Зебрафиш: визуализация Excel",
        layout="wide",
    )
    st.markdown(
        """
        <style>
          .block-container { padding-top: 1.4rem; }
          div[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
          label, p, span, div, input, textarea { letter-spacing: 0 !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title("Зебрафиш: визуализация из Excel")
    uploaded_file = st.sidebar.file_uploader("Загрузите файл .xlsx", type=["xlsx"])
    file_bytes, file_name = source_bytes(uploaded_file)
    token = hashlib.md5(file_bytes).hexdigest()[:10]

    try:
        parsed = parse_workbook(file_bytes)
    except Exception as exc:
        st.error(f"Не удалось прочитать файл: {exc}")
        st.stop()

    st.sidebar.caption(f"Источник: {file_name}")
    labels = sidebar_text_labels(default_labels(), token)
    colors = sidebar_colors(token)
    layout = sidebar_layout_controls(token)

    st.caption(
        f"Лист: `{parsed['sheet']}`. Строка с препаратами: `{parsed['header_row']}`. "
        "Смешанные значения вроде `↓↑` или `↓↑↓↑` автоматически окрашиваются под положением каждой стрелки; `≈` отображается как `-`."
    )

    edit_tabs = st.tabs([group["label"] for group in parsed["groups"]])
    edited_groups: list[dict[str, Any]] = []
    for tab, group in zip(edit_tabs, parsed["groups"]):
        with tab:
            edited_groups.append(edit_group(group, labels, token))

    st.divider()
    st.subheader("Визуализация")
    png_bytes = render_png_visualization(edited_groups, labels, colors, layout=layout, dpi=600)
    preview_size = Image.open(BytesIO(png_bytes)).size
    st.caption(
        f"PNG-превью готовой картинки: {preview_size[0]} x {preview_size[1]} px, "
        f"первый столбец {layout['row_width']} px, колонки препаратов {layout['cell_width']} px."
    )
    st.image(png_bytes, use_container_width=True)
    st.download_button(
        "Скачать PNG 600 dpi",
        data=png_bytes,
        file_name="zebrafish_visualization_600dpi.png",
        mime="image/png",
        type="primary",
    )


if __name__ == "__main__":
    main()
